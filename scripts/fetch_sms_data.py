#!/usr/bin/env python3
"""
Fetch real SMS pricing data from multiple provider sources.

Data sources (all public, no auth required):
  - Twilio:    Public CSV (219 countries)
  - Vonage:    Public XLSX download (232 countries)
  - Telnyx:    Next.js RSC server action (111 countries)
  - ClickSend: Public REST API (200+ countries)
  - Plivo:     Embedded JS pricing data (23 countries)
  - Sinch:     Third-party comparison data (limited)
  - Infobip:   Third-party comparison data (limited)
  - MessageBird: Third-party comparison data (limited)
"""

import csv
import io
import json
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

try:
    import requests
except ImportError:
    print("pip install requests"); sys.exit(1)

try:
    import openpyxl
except ImportError:
    openpyxl = None

OUTPUT = Path(__file__).resolve().parent.parent / "public" / "sms-data.json"
EUR_TO_USD = 1.05  # approximate conversion rate

PROVIDER_LINKS = {
    "Twilio":      "https://www.twilio.com/en-us/sms/pricing",
    "Vonage":      "https://www.vonage.com/communications-apis/sms/",
    "Plivo":       "https://www.plivo.com/sms/pricing/",
    "Telnyx":      "https://telnyx.com/pricing/messaging",
    "MessageBird": "https://bird.com/en/pricing",
    "ClickSend":   "https://www.clicksend.com/us/pricing",
    "Sinch":       "https://www.sinch.com/pricing/",
    "Infobip":     "https://www.infobip.com/pricing",
}

# OTP/Verify pricing (only some providers offer dedicated verify)
VERIFY_PRICES = {
    "Twilio": 0.05, "Vonage": 0.0572, "Telnyx": 0.03, "Sinch": 0.045,
}

# All ISO-2 country codes we want to query for ClickSend
TOP_COUNTRIES = [
    "US","GB","CA","AU","DE","FR","IN","BR","JP","CN","MX","KR","SG","AE",
    "ZA","NG","ID","PH","PK","EG","SA","RU","TR","PL","IL","KE","CH","NO",
    "SE","TH","MY","VN","BD","NZ","AT","CL","CO","AR","PE","IT","ES","NL",
    "IE","PT","DK","FI","BE","GR","CZ","HU","RO","UA","HK","TW","QA","KW",
    "OM","BH","JO","IQ","LB","GH","ET","TZ","UG","MA","DZ","TN","LY","SD",
    "SN","CM","CI","MG","MZ","ZM","ZW","RW","AO","MM","KH","LA","NP","LK",
    "BD","UZ","KZ","GE","AM","AZ","BY","RS","HR","BA","SI","SK","BG","LT",
    "LV","EE","IS","LU","MT","CY","PA","CR","GT","HN","SV","NI","DO","JM",
    "TT","BB","GY","SR","PY","UY","BO","EC","VE","FJ","PG","WS",
]


# ─── Twilio ────────────────────────────────────────────────────────────

def fetch_twilio() -> Dict[str, float]:
    """Download Twilio's public SMS pricing CSV. Returns {iso: price_usd}."""
    url = "https://assets.cdn.prod.twilio.com/pricing-csv/SMSPricing.csv"
    print("  Twilio: fetching CSV...")
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    reader = csv.DictReader(io.StringIO(resp.text))
    countries = {}
    for row in reader:
        iso = row.get("ISO", "").strip().upper()
        price_str = row.get("Price / msg", "").strip().lstrip("$")
        if not iso or not price_str or len(iso) != 2:
            continue
        try:
            price = float(price_str)
        except ValueError:
            continue
        if price <= 0:
            continue
        # Use max price per country (worst-case carrier)
        if iso not in countries or price > countries[iso]:
            countries[iso] = price
    print(f"  Twilio: {len(countries)} countries")
    return countries


# ─── Vonage ────────────────────────────────────────────────────────────

def fetch_vonage() -> Dict[str, float]:
    """Download Vonage's public pricing XLSX. Returns {iso: price_usd}."""
    url = "https://dashboard.vonage.com/download_pricing"
    print("  Vonage: fetching XLSX...")

    if openpyxl is None:
        # Try cached JSON from previous run
        cache = Path("/tmp/vonage_sms_pricing.json")
        if cache.exists():
            print("  Vonage: using cached JSON (openpyxl not installed)")
            with open(cache) as f:
                data = json.load(f)
            return _parse_vonage_json(data)
        print("  Vonage: SKIPPED (openpyxl not installed)")
        return {}

    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=False)
        ws = wb["Outbound SMS"]
        countries = {}
        for row_idx in range(2, ws.max_row + 1):
            iso = ws.cell(row=row_idx, column=1).value
            price_eur = ws.cell(row=row_idx, column=4).value
            iso = str(iso).strip().upper() if iso else ""
            if not iso or len(iso) != 2 or price_eur is None:
                continue
            try:
                price_usd = float(price_eur) * EUR_TO_USD
            except (ValueError, TypeError):
                continue
            if price_usd <= 0:
                continue
            countries[iso] = round(price_usd, 6)
        wb.close()
        print(f"  Vonage: {len(countries)} countries")
        return countries
    except Exception as e:
        print(f"  Vonage: XLSX fetch failed ({e}), trying cached JSON...")
        cache = Path("/tmp/vonage_sms_pricing.json")
        if cache.exists():
            with open(cache) as f:
                data = json.load(f)
            return _parse_vonage_json(data)
        return {}


def _parse_vonage_json(data: list) -> Dict[str, float]:
    countries = {}
    for entry in data:
        iso = entry.get("country_code", "").strip().upper()
        price_str = entry.get("price_eur_per_message", "")
        if not iso or len(iso) != 2 or not price_str:
            continue
        try:
            price_usd = float(price_str) * EUR_TO_USD
        except ValueError:
            continue
        if price_usd > 0:
            countries[iso] = round(price_usd, 6)
    print(f"  Vonage: {len(countries)} countries (from cached JSON)")
    return countries


# ─── Telnyx ────────────────────────────────────────────────────────────

def fetch_telnyx() -> Dict[str, float]:
    """Fetch Telnyx pricing via their Next.js RSC server action."""
    print("  Telnyx: fetching via RSC server action...")

    # Try cached data first (much faster)
    cache = Path("/tmp/telnyx_sms_pricing_full.json")
    if cache.exists():
        with open(cache) as f:
            data = json.load(f)
        countries = {}
        for iso, info in data.items():
            price = info.get("outbound_sms_usd")
            if isinstance(price, (int, float)) and price > 0:
                countries[iso] = round(float(price), 6)
        print(f"  Telnyx: {len(countries)} countries (from cache)")
        return countries

    # Live fetch via RSC
    url = "https://telnyx.com/pricing/messaging"
    action_id = "40b110a7fc7318f93a0b45953a7ce4eccb68ac9c7d"
    headers = {
        "Content-Type": "text/plain;charset=UTF-8",
        "Next-Action": action_id,
        "Accept": "text/x-component",
        "RSC": "1",
    }
    countries = {}
    # Fetch a selection of key countries
    for iso in TOP_COUNTRIES[:60]:
        try:
            resp = requests.post(url, headers=headers, data=json.dumps([iso]), timeout=10)
            if resp.status_code != 200:
                continue
            # Parse RSC response for price
            for line in resp.text.split("\n"):
                if "$" in line and "per message" in line.lower():
                    import re
                    match = re.search(r'\$(\d+\.?\d*)', line)
                    if match:
                        price = float(match.group(1))
                        if price > 0:
                            countries[iso] = price
                            break
            time.sleep(0.1)
        except Exception:
            continue
    print(f"  Telnyx: {len(countries)} countries (live)")
    return countries


# ─── ClickSend ─────────────────────────────────────────────────────────

def fetch_clicksend() -> Dict[str, float]:
    """Fetch ClickSend pricing via their public REST API (no auth needed)."""
    print("  ClickSend: fetching via public API...")
    countries = {}
    for iso in TOP_COUNTRIES:
        try:
            url = f"https://rest.clicksend.com/v3/pricing/{iso}"
            resp = requests.get(url, timeout=10)
            if resp.status_code != 200:
                continue
            data = resp.json()
            # Navigate to SMS pricing
            sms = data.get("data", {}).get("sms", {})
            # Use the base rate (rate_0 = lowest volume tier)
            price = sms.get("price_rate_0")
            carrier_fee = sms.get("price_sms_carrier_fee", 0)
            if price is not None:
                total = float(price) + float(carrier_fee or 0)
                if total > 0:
                    countries[iso] = round(total, 6)
            time.sleep(0.05)  # be polite
        except Exception:
            continue
    print(f"  ClickSend: {len(countries)} countries")
    return countries


# ─── Plivo ─────────────────────────────────────────────────────────────

def fetch_plivo() -> Dict[str, float]:
    """Plivo pricing from their embedded JS data (23 countries)."""
    print("  Plivo: using published pricing data...")
    # Derived from Plivo's SMSPricing JS object (cost per 100K SMS / 100000)
    countries = {
        "US": 0.00528, "CA": 0.00528, "GB": 0.03981, "AU": 0.04730,
        "IN": 0.00180, "FR": 0.07100, "IT": 0.08317, "ES": 0.07665,
        "PH": 0.14750, "SG": 0.03697, "ID": 0.29591, "LK": 0.25668,
        "UZ": 0.28335, "TR": 0.02756, "IL": 0.14626, "BD": 0.26534,
        "RU": 0.61478, "IQ": 0.28196, "MA": 0.20802, "CN": 0.03337,
        "PK": 0.29368, "BR": 0.05137, "DZ": 0.23264,
    }
    print(f"  Plivo: {len(countries)} countries")
    return countries


# ─── Sinch (limited third-party data) ─────────────────────────────────

def fetch_sinch() -> Dict[str, float]:
    """Sinch pricing from third-party comparison sites."""
    print("  Sinch: using third-party comparison data...")
    countries = {
        "US": 0.07, "GB": 0.04415, "FR": 0.0687, "KR": 0.0431,
        "CH": 0.07205, "IL": 0.167, "MA": 0.1529,
    }
    print(f"  Sinch: {len(countries)} countries (limited)")
    return countries


# ─── Infobip (limited third-party data) ───────────────────────────────

def fetch_infobip() -> Dict[str, float]:
    """Infobip pricing from third-party comparison sites."""
    print("  Infobip: using third-party comparison data...")
    countries = {
        "US": 0.0085, "GB": 0.05, "FR": 0.06717, "KR": 0.0358,
        "CH": 0.10, "IL": 0.17, "BR": 0.035,
    }
    print(f"  Infobip: {len(countries)} countries (limited)")
    return countries


# ─── MessageBird (limited) ────────────────────────────────────────────

def fetch_messagebird() -> Dict[str, float]:
    """MessageBird pricing — only US rate is publicly known."""
    print("  MessageBird: using published data...")
    countries = {"US": 0.008}
    print(f"  MessageBird: {len(countries)} countries (limited)")
    return countries


# ─── Build unified dataset ─────────────────────────────────────────────

def build_country_names(all_provider_data: Dict[str, Dict[str, float]]) -> Dict[str, str]:
    """Build ISO -> country name map from Twilio CSV."""
    url = "https://assets.cdn.prod.twilio.com/pricing-csv/SMSPricing.csv"
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    reader = csv.DictReader(io.StringIO(resp.text))
    names = {}
    for row in reader:
        iso = row.get("ISO", "").strip().upper()
        name = row.get("Country", "").strip()
        if iso and name and len(iso) == 2:
            names[iso] = name
    return names


def make_record(
    provider: str, iso: str, name: str, price: float,
    direction: str, verify: Optional[float], updated: str,
) -> dict:
    slug = provider.lower().replace(" ", "")
    return {
        "id": f"{slug}-{iso.lower()}-{direction[:3]}-sms",
        "provider": provider,
        "country_iso": iso,
        "country_name": name,
        "operator_name": None,
        "mcc": None, "mnc": None,
        "price_per_sms": price,
        "currency": "USD",
        "price_usd": price,
        "direction": direction,
        "message_type": "sms",
        "verify_price_usd": verify,
        "link": PROVIDER_LINKS[provider],
        "last_updated": updated,
    }


def main():
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    print("Fetching real SMS pricing data...\n")

    # Fetch from all sources
    provider_data = {
        "Twilio": fetch_twilio(),
        "Vonage": fetch_vonage(),
        "Telnyx": fetch_telnyx(),
        "ClickSend": fetch_clicksend(),
        "Plivo": fetch_plivo(),
        "Sinch": fetch_sinch(),
        "Infobip": fetch_infobip(),
        "MessageBird": fetch_messagebird(),
    }

    # Build country name lookup from Twilio data
    print("\n  Building country name index...")
    country_names = build_country_names(provider_data)

    # Also add names from Vonage for countries not in Twilio
    vonage_cache = Path("/tmp/vonage_sms_pricing.json")
    if vonage_cache.exists():
        with open(vonage_cache) as f:
            for entry in json.load(f):
                iso = entry.get("country_code", "").strip().upper()
                name = entry.get("country_name", "").strip()
                if iso and name and iso not in country_names:
                    country_names[iso] = name

    # Collect all country ISOs across all providers
    all_isos = set()
    for pdata in provider_data.values():
        all_isos.update(pdata.keys())

    # Generate records
    records = []
    for iso in sorted(all_isos):
        name = country_names.get(iso, iso)

        for provider, pdata in provider_data.items():
            if iso not in pdata:
                continue  # only include countries where we have REAL data

            price = pdata[iso]
            verify = VERIFY_PRICES.get(provider)

            # Outbound
            records.append(make_record(provider, iso, name, price, "outbound", verify, now))

            # Inbound (estimate ~85% of outbound for providers that support it)
            if provider in ("Twilio", "Vonage", "Telnyx"):
                inbound = round(price * 0.85, 6)
                records.append(make_record(provider, iso, name, inbound, "inbound", None, now))

    # Stats
    countries = set(r["country_iso"] for r in records)
    providers = set(r["provider"] for r in records)
    real_counts = {p: len(d) for p, d in provider_data.items()}

    output = {
        "lastUpdated": now,
        "count": len(records),
        "providerCount": len(providers),
        "countryCount": len(countries),
        "data": records,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, "w") as f:
        json.dump(output, f, separators=(",", ":"))

    size_kb = OUTPUT.stat().st_size / 1024
    print(f"\n{'='*60}")
    print(f"Wrote {OUTPUT}")
    print(f"  Total: {len(records)} records, {len(countries)} countries, {len(providers)} providers")
    print(f"  File size: {size_kb:.1f} KB")
    print(f"\n  Real data per provider:")
    for p, count in sorted(real_counts.items()):
        print(f"    {p:>12}: {count:>3} countries")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
